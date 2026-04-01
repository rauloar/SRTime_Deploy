from PySide6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QFileDialog, QPushButton
from models.attendance import AttendanceLog
from models.employee import User
import pandas as pd

class MovementsTab(QWidget):
    def __init__(self, session):
        super().__init__()
        self.session = session

        from PySide6.QtWidgets import QHBoxLayout, QLabel, QComboBox, QDateEdit, QCheckBox
        from PySide6.QtCore import QDate
        layout = QVBoxLayout()

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Fecha inicio:"))
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDate(QDate.currentDate())
        filter_layout.addWidget(self.date_start)

        filter_layout.addWidget(QLabel("Fecha fin:"))
        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDate(QDate.currentDate())
        filter_layout.addWidget(self.date_end)

        filter_layout.addWidget(QLabel("Usuario(s):"))
        self.user_selector = QComboBox()
        self.user_selector.setEditable(True)
        self.user_selector.addItem("Todos")
        users = self.session.query(User).filter(User.is_active==True).all()
        for user in users:
            display = (user.first_name or "") + " " + (user.last_name or "")
            self.user_selector.addItem(f"{user.identifier} - {display.strip()}", user.identifier)
        filter_layout.addWidget(self.user_selector)

        filter_layout.addWidget(QLabel("Vista:"))
        self.view_selector = QComboBox()
        self.view_selector.addItem("Vertical (Entradas/Salidas)", "vertical")
        self.view_selector.addItem("Horizontal (Por Día)", "horizontal")
        self.view_selector.currentIndexChanged.connect(self.load_data)
        filter_layout.addWidget(self.view_selector)

        layout.addLayout(filter_layout)

        from PySide6.QtWidgets import QSizePolicy
        from PySide6.QtWidgets import QHeaderView
        from PySide6.QtCore import Qt
        self.table = QTableWidget(self)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        layout.addWidget(self.table, stretch=1)

        from PySide6.QtWidgets import QStyle
        style = self.style()
        self.btn_refresh = QPushButton("Actualizar")
        self.btn_refresh.setIcon(style.standardIcon(QStyle.StandardPixmap.SP_BrowserReload))
        self.btn_refresh.clicked.connect(self.load_data)
        layout.addWidget(self.btn_refresh)

        self.btn_export = QPushButton("Exportar CSV")
        self.btn_export.setIcon(style.standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.btn_export.clicked.connect(self.export_csv)
        layout.addWidget(self.btn_export)

        self.btn_export_excel = QPushButton("Exportar Excel")
        self.btn_export_excel.setIcon(style.standardIcon(QStyle.StandardPixmap.SP_DriveHDIcon))
        self.btn_export_excel.clicked.connect(self.export_excel)
        layout.addWidget(self.btn_export_excel)

        self.setLayout(layout)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def update_user_selector(self):
        self.user_selector.clear()
        self.user_selector.addItem("Todos")
        users = self.session.query(User).filter(User.is_active==True).all()
        for user in users:
            display = (user.first_name or "") + " " + (user.last_name or "")
            self.user_selector.addItem(f"{user.identifier} - {display.strip()}", user.identifier)

    def load_data(self):
        start = self.date_start.date().toPython()
        end = self.date_end.date().toPython()
        user_value = self.user_selector.currentData()
        query = self.session.query(AttendanceLog, User).join(User, AttendanceLog.raw_identifier == User.identifier)
        query = query.filter(User.is_active == True)
        # Solo filtrar por usuario si no es 'Todos'
        if user_value and user_value != None and self.user_selector.currentText() != "Todos":
            query = query.filter(User.identifier == user_value)
        if start:
            query = query.filter(AttendanceLog.date >= start)
        if end:
            query = query.filter(AttendanceLog.date <= end)
        logs = query.all()
        
        self.table.setSortingEnabled(False)
        self.table.clearContents()
        
        view_mode = self.view_selector.currentData()
        
        if view_mode == "vertical":
            self.table.setRowCount(len(logs))
            self.table.setColumnCount(5)
            self.table.setHorizontalHeaderLabels(["Identificador", "Nombre", "Fecha", "Hora", "Movimiento"])
            for row, (log, user) in enumerate(logs):
                movimiento = "ENTRADA" if log.mark_type == 0 else "SALIDA"
                if user.first_name or user.last_name:
                    nombre = (user.first_name or "") + " " + (user.last_name or "")
                else:
                    nombre = log.raw_identifier
                fecha_str = log.date.strftime("%d/%m/%Y") if log.date else ""
                hora_str = log.time.strftime("%H:%M") if log.time else ""
                self.table.setItem(row, 0, QTableWidgetItem(log.raw_identifier))
                self.table.setItem(row, 1, QTableWidgetItem(nombre))
                self.table.setItem(row, 2, QTableWidgetItem(fecha_str))
                self.table.setItem(row, 3, QTableWidgetItem(hora_str))
                self.table.setItem(row, 4, QTableWidgetItem(movimiento))
        else:
            # Vista horizontal agrupando por usuario y fecha
            from collections import defaultdict
            grouped = defaultdict(list)
            for log, user in logs:
                grouped[(log.raw_identifier, log.date)].append((log, user))
                
            self.table.setRowCount(len(grouped))
            self.table.setColumnCount(5)
            self.table.setHorizontalHeaderLabels(["Identificador", "Nombre", "Fecha", "Entrada", "Salida"])
            
            row = 0
            for (identifier, date), items in grouped.items():
                items.sort(key=lambda x: x[0].time)
                user = items[0][1]
                nombre = (user.first_name or "") + " " + (user.last_name or "") if user.first_name or user.last_name else identifier
                fecha_str = date.strftime("%d/%m/%Y") if date else ""
                
                in_time_str = ""
                out_time_str = ""
                
                in_logs = [x for x in items if x[0].mark_type == 0]
                out_logs = [x for x in items if x[0].mark_type == 1]
                
                if in_logs and out_logs:
                    in_time_str = in_logs[0][0].time.strftime("%H:%M")
                    out_time_str = out_logs[-1][0].time.strftime("%H:%M")
                elif len(items) >= 2:
                    in_time_str = items[0][0].time.strftime("%H:%M")
                    out_time_str = items[-1][0].time.strftime("%H:%M")
                elif len(items) == 1:
                    log_item = items[0][0]
                    if log_item.mark_type == 0:
                        in_time_str = log_item.time.strftime("%H:%M")
                    elif log_item.mark_type == 1:
                        out_time_str = log_item.time.strftime("%H:%M")
                    else:
                        in_time_str = log_item.time.strftime("%H:%M")
                        
                self.table.setItem(row, 0, QTableWidgetItem(identifier))
                self.table.setItem(row, 1, QTableWidgetItem(nombre))
                self.table.setItem(row, 2, QTableWidgetItem(fecha_str))
                self.table.setItem(row, 3, QTableWidgetItem(in_time_str))
                self.table.setItem(row, 4, QTableWidgetItem(out_time_str))
                row += 1

        self.table.setSortingEnabled(True)

    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Guardar CSV", "", "CSV Files (*.csv)")
        if path:
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            data = []
            
            selected_indexes = self.table.selectionModel().selectedRows()
            selected_rows = [i.row() for i in selected_indexes]
            rows_to_export = selected_rows if selected_rows else range(self.table.rowCount())

            for row in rows_to_export:
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            df = pd.DataFrame(data, columns=headers)
            df.to_csv(path, index=False)

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "", "Excel Files (*.xlsx)")
        if path:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            data = []
            
            selected_indexes = self.table.selectionModel().selectedRows()
            selected_rows = [i.row() for i in selected_indexes]
            rows_to_export = selected_rows if selected_rows else range(self.table.rowCount())

            for row in rows_to_export:
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Crear workbook y hoja
            wb = openpyxl.Workbook()
            ws = wb.active
            import os
            sheet_name = os.path.splitext(os.path.basename(path))[0]
            ws.title = sheet_name
            # Encabezados en negrita
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Datos
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max(15, max_length + 2)
            wb.save(path)